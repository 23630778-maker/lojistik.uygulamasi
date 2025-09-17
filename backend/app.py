from flask import Flask, render_template, request, redirect, url_for, flash
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import shutil  # Kopyalama için gerekli

app = Flask(__name__)
app.secret_key = "supersecretkey"

# Excel dosyalarının yolları
EXCEL_FILE_LOCAL = r"C:\Users\nisak\Desktop\lojistik.xlsx"
EXCEL_FILE_ONEDRIVE = r"C:\Users\nisak\OneDrive\lojistik.xlsx"

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        try:
            # Form verilerini al
            tarih = request.form.get("tarih") or datetime.now().strftime("%Y-%m-%d")
            iscikissaat = request.form.get("iscikissaat") or "00:00"
            plaka = request.form.get("plaka")
            cikiskm = float(request.form.get("cikiskm") or 0)
            kumgirissaat = request.form.get("kumgirissaat") or "00:00"
            giriskm = float(request.form.get("giriskm") or 0)
            kumcikissaat = request.form.get("kumcikissaat") or "00:00"
            isletmegiriskm = float(request.form.get("isletmegiriskm") or 0)
            isletmegirissaat = request.form.get("isletmegirissaat") or "00:00"
            farkkm = giriskm - cikiskm
            uretici = request.form.get("uretici")
            ureticikm = float(request.form.get("ureticikm") or 0)
            tonaj = int(request.form.get("tonaj") or 0)

            # Excel dosyası var mı kontrol et (önce lokal)
            if os.path.exists(EXCEL_FILE_LOCAL):
                wb = load_workbook(EXCEL_FILE_LOCAL)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # Başlıkları ekle
                ws.append([
                    "tarih", "iscikissaat", "plaka", "cikiskm", "kumgirissaat",
                    "giriskm", "kumcikissaat", "isletmegiriskm", "isletmegirissaat",
                    "farkkm", "uretici", "ureticikm", "tonaj"
                ])

            # Yeni veriyi ekle
            ws.append([
                tarih, iscikissaat, plaka, cikiskm, kumgirissaat,
                giriskm, kumcikissaat, isletmegiriskm, isletmegirissaat,
                farkkm, uretici, ureticikm, tonaj
            ])

            # Önce lokal kaydet
            wb.save(EXCEL_FILE_LOCAL)

            # Sonra OneDrive klasörüne kopyala
            shutil.copy(EXCEL_FILE_LOCAL, EXCEL_FILE_ONEDRIVE)

            flash("Kayıt başarıyla eklendi ve OneDrive’a kaydedildi!", "success")
            return redirect(url_for("form"))

        except Exception as e:
            flash(f"Hata oluştu: {e}", "danger")
            return redirect(url_for("form"))

    return render_template("form.html")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
